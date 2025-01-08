import os
import re
import statistics
from openpyxl import load_workbook

'''
xlsx结果文件工具
'''

def get_xlsx_files(log_path):
    # 遍历目录，查找以result_worker开头且以.xlsx结尾的文件
    xlsx_files = [os.path.join(log_path, file)
                  for file in os.listdir(log_path)
                  if file[0].isdigit() and file.endswith('.xlsx')]
    # 对所有xlsx文件按照后缀index进行排序
    sorted_files = sorted(xlsx_files, key=lambda x: int(re.search(r'(\d+)\.xlsx$', x).group(1)))
    return sorted_files

def get_all_file_columns(files, column_index):
    result = []
    for file in files:
        workbook = load_workbook(file, data_only=True)  # 打开 Excel 文件
        sheet = workbook.active  # 默认读取活动工作表
        # 遍历指定列，从第二行开始读取数据
        for row in sheet.iter_rows(min_row=2, min_col=column_index + 1, max_col=column_index + 1, values_only=True):
            value = row[0]  # 获取单元格值
            if value is not None:  # 忽略空值
                result.append(value)
    return result

def weighted_sum(weight_list, result_list):
    # 计算权重总和
    total_weight = sum(weight_list)
    # 计算加权和
    weighted_total = sum((w / total_weight) * r for w, r in zip(weight_list, result_list))
    return weighted_total

def merge_result0(log_fp):
    files = get_xlsx_files(log_fp)

    weight_list = get_all_file_columns(files, column_index=2)
    cloud_list = get_all_file_columns(files, column_index=4)
    local_list = get_all_file_columns(files, column_index=5)
    local_plus_list = get_all_file_columns(files, column_index=6)
    mpda_minus_list = get_all_file_columns(files, column_index=7)
    mpda_list = get_all_file_columns(files, column_index=8)

    print(f'Cloud = {weighted_sum(weight_list, cloud_list)}')
    print(f'Local= {weighted_sum(weight_list, local_list)}')
    print(f'Local+ = {weighted_sum(weight_list, local_plus_list)}')
    print(f'MPDA- = {weighted_sum(weight_list, mpda_minus_list)}')
    print(f'MPDA = {weighted_sum(weight_list, mpda_list)}')

def merge_result(log_fp):
    files = get_xlsx_files(log_fp)

    weight_list = get_all_file_columns(files, column_index=2)
    local_plus_list = get_all_file_columns(files, column_index=4)
    mpda_minus_list = get_all_file_columns(files, column_index=5)
    mpda_list = get_all_file_columns(files, column_index=6)

    print(f'Local+ = {weighted_sum(weight_list, local_plus_list)}')
    print(f'MPDA- = {weighted_sum(weight_list, mpda_minus_list)}')
    print(f'MPDA = {weighted_sum(weight_list, mpda_list)}')

def merge_local_result(log_fp):
    files = get_xlsx_files(log_fp)
    weight_list = get_all_file_columns(files, column_index=1)
    local_list = get_all_file_columns(files, column_index=3)
    print(f'Local= {weighted_sum(weight_list, local_list)}')

def merge_cloud_result(log_fp):
    files = get_xlsx_files(log_fp)
    weight_list = get_all_file_columns(files, column_index=1)
    cloud_list = get_all_file_columns(files, column_index=3)
    print(f'Cloud= {weighted_sum(weight_list, cloud_list)}')

log_fp = '/home/chao/workspace/MPDA-implementation/log/transfer_movielens_NCF_100_recall_item_pair_similarity'
merge_result(log_fp)
'''
Cloud = 0.6139675060163279
Local= 0.6141152495464451
Local+ = 0.6145060692948374
MPDA- = 0.6227206597601547
MPDA = 0.6193539726434936
'''
